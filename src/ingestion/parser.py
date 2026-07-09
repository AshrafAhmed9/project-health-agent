import openpyxl
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List

class ExcelParser:
    """
    Parses S2P Project.xlsx and Project Plan B.xlsx.
    Determines schema type automatically and extracts summary and tasks data.
    """
    
    def detect_schema(self, ws) -> str:
        """
        Detects schema type from first row headers.
        's2p' or 'plan_b'
        """
        # Read the first row headers
        first_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not first_row:
            raise ValueError("Empty sheet or no headers found")
            
        if first_row[0] == "Project Name":
            return "s2p"
        elif first_row[0] == "No.of days Until Today":
            return "plan_b"
        else:
            # Check other signature columns
            headers_str = "".join([str(h) for h in first_row if h])
            if "Variance2" in headers_str or "Baseline Start2" in headers_str:
                return "plan_b"
            elif "Outokumpu" in headers_str or "Titan" in headers_str or "Ancestors" in headers_str:
                return "s2p"
            return "s2p" # Default fallback
            
    def parse_summary_sheet(self, ws) -> Dict[str, Any]:
        """
        Parses the 'Summary' sheet.
        """
        data = {}
        for row in ws.iter_rows(values_only=True):
            if len(row) >= 2 and row[0] is not None:
                key = str(row[0]).strip()
                val = row[1]
                
                # Clean value
                if isinstance(val, str) and val.startswith("#"):
                    val = None
                data[key] = val
        return data

    def parse_comments_sheet(self, ws) -> List[Dict[str, Any]]:
        """
        Parses the 'Comments' sheet if it exists.
        """
        comments = []
        # Header is usually row 1: Row Reference, Comment, Author, Timestamp
        headers = []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return comments
            
        first_row = rows[0]
        if first_row and len(first_row) >= 1:
            headers = [str(h).strip() if h is not None else f"Col_{i}" for i, h in enumerate(first_row)]
            
        for row in rows[1:]:
            # Skip completely empty rows
            if not any(row):
                continue
            comment_item = {}
            for idx, cell_val in enumerate(row):
                if idx < len(headers):
                    key = headers[idx]
                    val = cell_val
                    if isinstance(val, str) and val.startswith("#"):
                        val = None
                    comment_item[key] = val
            comments.append(comment_item)
        return comments

    def parse(self, filepath: Path) -> Dict[str, Any]:
        """
        Reads excel file and parses all relevant sheets.
        """
        if not filepath.exists():
            raise FileNotFoundError(f"File not found: {filepath}")
            
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet_names = wb.sheetnames
        
        # Summary parsing
        summary_data = {}
        if "Summary" in sheet_names:
            summary_data = self.parse_summary_sheet(wb["Summary"])
            
        # Comments parsing
        comments_data = []
        if "Comments" in sheet_names:
            comments_data = self.parse_comments_sheet(wb["Comments"])
            
        # Find main tasks sheet
        # Usually first sheet, or one named like "Project Plan" or contains tasks
        main_sheet_name = None
        for name in sheet_names:
            if name != "Summary" and name != "Comments":
                main_sheet_name = name
                break
        if not main_sheet_name:
            main_sheet_name = sheet_names[0]
            
        ws_main = wb[main_sheet_name]
        schema_type = self.detect_schema(ws_main)
        
        # Parse main tasks
        rows = list(ws_main.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"No task rows in sheet {main_sheet_name}")
            
        headers = [str(cell).strip() if cell is not None else f"Col_{i}" for i, cell in enumerate(rows[0])]
        
        tasks_raw = []
        for row_idx, row in enumerate(rows[1:], start=2):
            # Skip empty rows
            if not any(row):
                continue
            
            task_dict = {"_row_num": row_idx}
            for col_idx, cell_val in enumerate(row):
                if col_idx < len(headers):
                    key = headers[col_idx]
                    val = cell_val
                    # Filter out unparseable formulas
                    if isinstance(val, str) and val.startswith("#"):
                        val = None
                    task_dict[key] = val
            tasks_raw.append(task_dict)
            
        return {
            "source_file": str(filepath.name),
            "schema_type": schema_type,
            "summary": summary_data,
            "comments": comments_data,
            "tasks_raw": tasks_raw,
            "headers": headers,
            "main_sheet_name": main_sheet_name
        }
